import sys
import os
import openpyxl

# 1. 获取输入参数 (从环境变量获取，这比 sys.argv 更稳妥)
# 如果你是用 GitHub Action 的 inputs 传参，最好用 os.environ
# 但为了兼容你原来的习惯，我写了两套获取方式
col_a = os.environ.get('INPUT_COL_A') or (sys.argv[1] if len(sys.argv) > 1 else "")
col_b = os.environ.get('INPUT_COL_B') or (sys.argv[2] if len(sys.argv) > 2 else "")
col_c = os.environ.get('INPUT_COL_C') or (sys.argv[3] if len(sys.argv) > 3 else "")
action = os.environ.get('INPUT_ACTION_TYPE', 'add') # 默认是 add

# 去除首尾空格，防止匹配失败
col_a = str(col_a).strip()
col_b = str(col_b).strip()
col_c = str(col_c).strip()

file_path = 'data.xlsx'
print(f"操作: {action}, 数据: [{col_a}, {col_b}, {col_c}]")

# 2. 加载 Excel 文件
if not os.path.exists(file_path):
    print("文件不存在，新建文件...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["POS名称", "设备地址", "二维码"]) # 写入表头
else:
    print("加载现有文件...")
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

# 3. 核心逻辑：遍历所有行，查找是否已存在
found = False
rows_to_delete = [] # 记录需要删除的行号（倒序删除，防止索引错乱）

# 从第2行开始遍历（跳过表头）
# ws.max_row 是最大行数
for row_idx in range(2, ws.max_row + 1):
    #以此类推：cell(row, 1)是第一列(POS名称)
    cell_val = ws.cell(row=row_idx, column=1).value
    
    # 强制转为字符串比较，解决数字格式不匹配问题
    if str(cell_val).strip() == col_a:
        found = True
        
        if action == 'delete':
            # 如果是删除操作，记录行号
            rows_to_delete.append(row_idx)
            print(f"找到待删除行: {row_idx}")
            
        else:
            # 如果是 update/add 操作，直接原地修改
            ws.cell(row=row_idx, column=2).value = col_b
            ws.cell(row=row_idx, column=3).value = col_c
            print(f"已更新行 {row_idx}: [{col_a}]")
            
            # 【关键】为了防止以后有重复数据，我们继续遍历，
            # 如果后面还有同名的行，全部标记删除，只保留这第一行！
            # (这是一个自我修复机制)
            for dup_idx in range(row_idx + 1, ws.max_row + 1):
                dup_val = ws.cell(row=dup_idx, column=1).value
                if str(dup_val).strip() == col_a:
                    rows_to_delete.append(dup_idx)
            break

# 4. 执行删除操作 (倒序删除，非常重要！)
if rows_to_delete:
    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r)
    print(f"已删除 {len(rows_to_delete)} 行旧数据。")

# 5. 如果没找到，并且不是删除操作 -> 新增到末尾
if not found and action != 'delete':
    ws.append([col_a, col_b, col_c])
    print(f"未找到 [{col_a}]，已新增到末尾。")

# 6. 保存文件
wb.save(file_path)
print("保存成功！")
